from django import forms
from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path

from fixtures.models import (
    Fixtures, FixtureCategory, FixtureCourt, FixtureTeam, FixtureMatch,
)
from tournament.models import Tournament


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def _parse_excel(excel_file):
    """
    Parse the fixture Excel workbook and return a structured list.

    Each sheet → one category.
    Within a sheet, rows whose column B (index 1) starts with "Court" mark the
    start of a new court block.  The next 4-5 rows contain team data and match
    pairs spread across two column groups (F-H and J-L, indices 5-7 / 9-11).
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    categories = []

    for sheet_order, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Pull the descriptive title from the first non-empty row (row index 1)
        title_text = ""
        for row in all_rows[1:4]:
            if row[1] and "Fixtures" in str(row[1]):
                title_text = str(row[1]).strip()
                break

        category_data = {
            "name": sheet_name.strip(),
            "title": title_text,
            "order": sheet_order,
            "courts": [],
        }

        current_court = None
        court_order = 0

        for row in all_rows:
            col_b = row[1]

            # Court header row: column B starts with "Court"
            if col_b and str(col_b).strip().startswith("Court"):
                current_court = {
                    "name": str(col_b).strip(),
                    "order": court_order,
                    "teams": [],
                    "matches": [],
                }
                category_data["courts"].append(current_court)
                court_order += 1
                continue

            # Team / match data rows
            if current_court is None:
                continue

            col_c, col_d = row[2], row[3]   # Team #, Pair name
            col_f, col_g, col_h = row[5], row[6], row[7]   # Matches 1-5
            col_j, col_k, col_l = row[9], row[10], row[11]  # Matches 6-10

            team_number = str(col_c).strip() if col_c else ""
            pair_name = str(col_d).strip() if col_d else ""

            if team_number and pair_name:
                current_court["teams"].append({
                    "team_number": team_number,
                    "pair_name": pair_name,
                    "order": len(current_court["teams"]),
                })

            if col_f is not None and col_g and col_h:
                current_court["matches"].append({
                    "match_number": int(col_f),
                    "team1": str(col_g).strip(),
                    "team2": str(col_h).strip(),
                })

            if col_j is not None and col_k and col_l:
                current_court["matches"].append({
                    "match_number": int(col_j),
                    "team1": str(col_k).strip(),
                    "team2": str(col_l).strip(),
                })

        categories.append(category_data)

    return categories


def _import_fixtures(excel_file, tournament=None):
    """Delete fixtures for the given tournament then import from *excel_file*."""
    if tournament:
        FixtureCategory.objects.filter(tournament=tournament).delete()
    else:
        FixtureCategory.objects.all().delete()

    categories = _parse_excel(excel_file)
    for cat_data in categories:
        category = FixtureCategory.objects.create(
            tournament=tournament,
            name=cat_data["name"],
            title=cat_data["title"],
            order=cat_data["order"],
        )
        for court_data in cat_data["courts"]:
            court = FixtureCourt.objects.create(
                category=category,
                name=court_data["name"],
                order=court_data["order"],
            )
            for t in court_data["teams"]:
                FixtureTeam.objects.create(
                    court=court,
                    team_number=t["team_number"],
                    pair_name=t["pair_name"],
                    order=t["order"],
                )
            for m in court_data["matches"]:
                FixtureMatch.objects.create(
                    court=court,
                    match_number=m["match_number"],
                    team1=m["team1"],
                    team2=m["team2"],
                )

    total_courts = sum(len(c["courts"]) for c in categories)
    return len(categories), total_courts


# ---------------------------------------------------------------------------
# Upload form
# ---------------------------------------------------------------------------

class FixtureUploadForm(forms.Form):
    tournament = forms.ModelChoiceField(
        queryset=Tournament.objects.order_by('-tournament_date_time'),
        label="Tournament",
        help_text="Fixtures will be linked to this tournament. "
                  "Existing fixture data for this tournament will be replaced.",
    )
    excel_file = forms.FileField(
        label="Excel file (.xlsx)",
        help_text="Upload the tournament fixtures Excel workbook.",
    )


# ---------------------------------------------------------------------------
# Admin registrations
# ---------------------------------------------------------------------------

class FixtureTeamInline(admin.TabularInline):
    model = FixtureTeam
    extra = 0


class FixtureMatchInline(admin.TabularInline):
    model = FixtureMatch
    extra = 0


@admin.register(FixtureCourt)
class FixtureCourtAdmin(admin.ModelAdmin):
    list_display = ["name", "category", "order"]
    list_filter = ["category"]
    inlines = [FixtureTeamInline, FixtureMatchInline]


class FixtureCourtInline(admin.TabularInline):
    model = FixtureCourt
    extra = 0
    show_change_link = True
    fields = ["name", "order"]


@admin.register(FixtureCategory)
class FixtureCategoryAdmin(admin.ModelAdmin):
    list_display = ["name", "tournament", "title", "order"]
    list_filter = ["tournament"]
    inlines = [FixtureCourtInline]

    # ---- custom "Upload Fixtures" page -----------------------------------

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                "upload-fixtures/",
                self.admin_site.admin_view(self.upload_fixtures_view),
                name="fixtures_fixturecategory_upload",
            ),
        ]
        return extra + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["upload_url"] = "upload-fixtures/"
        return super().changelist_view(request, extra_context=extra_context)

    def upload_fixtures_view(self, request):
        if request.method == "POST":
            form = FixtureUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    tournament = form.cleaned_data["tournament"]
                    n_cats, n_courts = _import_fixtures(request.FILES["excel_file"], tournament)
                    messages.success(
                        request,
                        f"Successfully imported {n_cats} categories and {n_courts} courts.",
                    )
                except Exception as exc:
                    messages.error(request, f"Import failed: {exc}")
                return redirect("..")
        else:
            form = FixtureUploadForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": "Upload Fixtures from Excel",
            "opts": self.model._meta,
        }
        return render(request, "admin/fixtures/upload_fixtures.html", context)
